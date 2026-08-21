
"""Модуль экспорта данных в Excel."""
import os
import subprocess
import platform
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils import today_str


class ExcelExporter:
    """Экспорт данных таблиц в Excel (.xlsx)."""

    @staticmethod
    def export_data(headers: list, rows: list, filename: str = None, sheet_title: str = "Данные"):
        """
        Экспортирует данные в Excel-файл.

        :param headers: заголовки столбцов (без галочки)
        :param rows:    список кортежей со значениями (без галочки)
        :param filename: имя файла (если None — генерируется)
        :param sheet_title: название листа
        :return: полный путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        # ── Стили ──────────────────────────────────────────
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # ── Заголовки ──────────────────────────────────────
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ── Данные ─────────────────────────────────────────
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # ── Автоширина ─────────────────────────────────────
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(rows) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_length = max(max_length, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

        # ── Сохранение ─────────────────────────────────────
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.xlsx"

        date_str = today_str()
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.exists(desktop):
            desktop = os.path.expanduser("~")
        save_dir = os.path.join(desktop, "Таблицы", date_str)
        os.makedirs(save_dir, exist_ok=True)
        filepath = os.path.join(save_dir, filename)

        wb.save(filepath)
        return filepath

    @staticmethod
    def open_file(filepath: str):
        """Открывает файл в системном приложении по умолчанию."""
        try:
            if platform.system() == 'Windows':
                os.startfile(filepath)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', filepath])
            else:
                subprocess.run(['xdg-open', filepath])
        except Exception as e:
            print(f"Не удалось открыть файл: {e}")
